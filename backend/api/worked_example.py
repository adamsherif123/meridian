"""Worked-example capture endpoint.

POST /api/v1/boards/{board_id}/worked-example  (multipart/form-data)
  Fields:
    email             UploadFile   — the example email (txt/pdf/any)
    attachments       UploadFile[] — N attachment files (invoices, COAs, etc.)
    expected_output   UploadFile   — expected results (CSV primary; xlsx best-effort)
    fixture_subject   str optional — shipment key to pass as email subject (e.g. MAWB); "" = auto

Behavior:
- Uploads email+attachments to Supabase storage at worked-examples/{board_id}/
- Replaces sample_files rows for this board (idempotent)
- Derives the expected output schema from the board (frozen spec → board graph → file headers)
- Maps the uploaded file onto that schema via an LLM so arbitrary column names, ordering,
  synonyms, and orientation all work
- Writes eval case JSON to backend/evals/cases/board_{board_id[:8]}_worked.json
- Returns {captured, case_path, answer_key, needs_confirmation, output_schema, schema_source}

needs_confirmation=True only when no schema is derivable AND the file is unreadable.
The agent is always buildable regardless of whether auto-grading succeeds.
"""

import csv
import io
import json
import logging
import os
import pathlib
import re
import uuid
from typing import List, Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from supabase import create_client

router = APIRouter(prefix="/api/v1/boards", tags=["worked-example"])
log = logging.getLogger(__name__)

BUCKET         = "sample-files"
STORAGE_PREFIX = "worked-examples"
TEXT_CAP       = 20_000
CASES_DIR      = pathlib.Path(__file__).parent.parent / "evals" / "cases"
_LLM_MODEL     = "claude-sonnet-4-6"
_LLM_MAX_TOK   = 1024


# ── Infrastructure helpers ─────────────────────────────────────────────────────

def _sb():
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_KEY", "")
    if not url or not key:
        raise HTTPException(503, "Supabase not configured")
    return create_client(url, key)


def _llm_client():
    """Return an Anthropic client, or None — never raises (LLM mapping is optional)."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None
    try:
        import anthropic
        return anthropic.Anthropic(api_key=api_key)
    except Exception:
        return None


def _coerce_value(raw: str):
    """Generic coercion: int → float → str.  Returns None for blank input."""
    s = (raw or "").strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


# ── Schema derivation ──────────────────────────────────────────────────────────

def _derive_output_schema(board_id: str, sb) -> tuple[list[str], str]:
    """Return (column_list, source) for this board's expected output.

    Precedence:
      1. frozen_specs  → spec.nodes → Report node's declared columns
      2. board_graphs  → nodes      → Report node's declared columns
      3. ([], "none")  — caller falls back to file headers or LLM inference
    """
    from backend.evals.consistency import _report_columns_from_spec  # reuse, don't duplicate

    # 1. frozen_spec (most authoritative — frozen before codegen)
    try:
        res = (
            sb.table("frozen_specs")
            .select("spec")
            .eq("board_id", board_id)
            .limit(1)
            .execute()
        )
        rows = (res.data or []) if res else []
        if rows and rows[0].get("spec"):
            cols = _report_columns_from_spec(rows[0]["spec"])
            if cols:
                log.info("_derive_output_schema: %d cols from frozen_spec", len(cols))
                return cols, "frozen_spec"
    except Exception as exc:
        log.warning("_derive_output_schema: frozen_spec lookup failed: %s", exc)

    # 2. live board_graphs (used when spec not yet frozen)
    try:
        res = (
            sb.table("board_graphs")
            .select("nodes")
            .eq("board_id", board_id)
            .limit(1)
            .execute()
        )
        rows = (res.data or []) if res else []
        if rows:
            nodes = rows[0].get("nodes") or []
            # reuse _report_columns_from_spec by wrapping the nodes list
            cols = _report_columns_from_spec({"nodes": nodes})
            if cols:
                log.info("_derive_output_schema: %d cols from board_graph", len(cols))
                return cols, "board_graph"
    except Exception as exc:
        log.warning("_derive_output_schema: board_graph lookup failed: %s", exc)

    log.info("_derive_output_schema: no schema found in DB for board %s", board_id)
    return [], "none"


# ── File text extraction ───────────────────────────────────────────────────────

def _extract_text(content: bytes, filename: str, mime: str) -> str:
    """Extract plain text from PDF, CSV, or TXT content."""
    name_lower = (filename or "").lower()
    if mime == "application/pdf" or name_lower.endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                return "\n".join(p.extract_text() or "" for p in pdf.pages)[:TEXT_CAP]
        except Exception as exc:
            log.warning("pdfplumber failed for %s: %s", filename, exc)
            return ""
    if mime in ("text/plain", "text/csv") or name_lower.endswith((".txt", ".csv")):
        try:
            return content.decode("utf-8", errors="replace")[:TEXT_CAP]
        except Exception:
            return ""
    return ""


def _xlsx_to_text(content: bytes) -> str:
    """Convert xlsx to a CSV-like text representation for LLM consumption."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        lines = []
        for row in rows[:50]:
            lines.append(",".join("" if v is None else str(v) for v in row))
        return "\n".join(lines)[:TEXT_CAP]
    except Exception as exc:
        log.warning("_xlsx_to_text failed: %s", exc)
        return ""


def _render_file_as_text(content: bytes, filename: str, mime: str) -> str:
    """Render the expected-output file as plain text for LLM mapping."""
    name_lower = (filename or "").lower()
    if name_lower.endswith((".xlsx", ".xls")):
        return _xlsx_to_text(content)
    text = _extract_text(content, filename, mime)
    if not text and mime == "text/plain":
        try:
            text = content.decode("utf-8", errors="replace")[:TEXT_CAP]
        except Exception:
            pass
    return text


# ── Fast-path deterministic parse ──────────────────────────────────────────────

def _fast_path_parse(
    content: bytes, filename: str, mime: str, schema_cols: list[str]
) -> dict | None:
    """Fill the answer key directly from file headers — no LLM.

    - If schema_cols is non-empty: returns a dict only when ALL schema_cols are
      found (case-insensitive) in the file's headers.  Missing columns → None
      (fall through to LLM).
    - If schema_cols is empty: reads the file's own headers and returns their
      values directly (the file IS the schema).
    - Returns None for non-CSV/XLSX files (let caller fall through to LLM).
    """
    name_lower = (filename or "").lower()
    is_csv  = name_lower.endswith(".csv") or mime in ("text/csv", "text/plain")
    is_xlsx = name_lower.endswith((".xlsx", ".xls"))

    if is_csv:
        try:
            text = content.decode("utf-8", errors="replace").strip()
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return None
            row_norm = {k.strip().lower(): v for k, v in rows[0].items()}
            if not schema_cols:
                return {k: _coerce_value(v) for k, v in row_norm.items()}
            result = {}
            for col in schema_cols:
                val = row_norm.get(col.strip().lower())
                if val is None:
                    return None   # missing — fall through to LLM
                result[col] = _coerce_value(val)
            return result
        except Exception:
            return None

    if is_xlsx:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return None
            headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
            values  = rows[1]
            row_map = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            if not schema_cols:
                return {k: _coerce_value(str(v) if v is not None else "") for k, v in row_map.items()}
            result = {}
            for col in schema_cols:
                v = row_map.get(col.strip().lower())
                if v is None:
                    return None   # missing — fall through to LLM
                result[col] = _coerce_value(str(v))
            return result
        except Exception:
            return None

    return None


# ── LLM mapping ────────────────────────────────────────────────────────────────

_LLM_SYSTEM = """\
You extract expected output values from a document that shows what an AI agent's report should look like.

Rules:
- Map by MEANING, not exact header text.  Tolerate synonyms, different casing, underscores vs spaces,
  and either row-oriented (header row + data row) or column-oriented (label | value pairs) layouts.
- Use null for a field only when its value is genuinely absent from the document.
- Output ONLY a valid JSON object — no prose, no markdown code fences, no explanation.
- For numeric fields, output the number as a JSON number (not a quoted string) when possible.
- If no target columns are provided, extract all key-value pairs you find as a flat JSON object.\
"""


def _llm_map_to_schema(
    schema_cols: list[str], file_text: str, board_id: str
) -> dict | None:
    """Ask Claude to map file_text onto schema_cols.

    Returns {col: value, ...} (values may be None) or None on total failure.
    Caller sets needs_confirmation=True when this returns None.
    """
    client = _llm_client()
    if client is None:
        log.warning("_llm_map_to_schema: ANTHROPIC_API_KEY not set — skipping LLM mapping")
        return None
    if not file_text.strip():
        log.warning("_llm_map_to_schema: empty file text — cannot map")
        return None

    if schema_cols:
        cols_block = "\n".join(f"- {c}" for c in schema_cols)
        user_msg = (
            f"Target output columns:\n{cols_block}\n\n"
            f"File content:\n{file_text[:TEXT_CAP]}"
        )
    else:
        user_msg = (
            "Extract all output fields as a flat JSON object.\n\n"
            f"File content:\n{file_text[:TEXT_CAP]}"
        )

    try:
        resp = client.messages.create(
            model=_LLM_MODEL,
            max_tokens=_LLM_MAX_TOK,
            system=_LLM_SYSTEM,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw = (resp.content[0].text or "").strip()
    except Exception as exc:
        log.warning("_llm_map_to_schema: LLM call failed: %s", exc)
        return None

    # Strip any markdown code fences the model may add despite the instructions
    raw = re.sub(r"^```[a-z]*\n?", "", raw).rstrip("`").strip()

    try:
        result = json.loads(raw)
        if not isinstance(result, dict):
            log.warning("_llm_map_to_schema: LLM returned %s, expected dict", type(result))
            return None
        log.info("_llm_map_to_schema: mapped %d fields for board %s", len(result), board_id)
        return result
    except json.JSONDecodeError as exc:
        log.warning("_llm_map_to_schema: JSON parse failed (%s) — raw=%r", exc, raw[:200])
        return None


# ── Main parse orchestrator ────────────────────────────────────────────────────

def _parse_expected_output(
    content: bytes,
    filename: str,
    mime: str,
    board_id: str,
    sb,
) -> tuple[dict, bool, list[str], str]:
    """Parse the expected-output file → (answer_key, needs_confirmation, schema_cols, schema_source).

    Schema derivation precedence:
      frozen_spec → board_graph Report node → file's own headers → LLM-inferred

    needs_confirmation is True only when no schema is derivable AND the file is unreadable.
    """
    # Step 1: derive schema from the board (spec or graph)
    schema_cols, schema_source = _derive_output_schema(board_id, sb)

    # Step 2: cheap deterministic fast-path (header-exact match)
    fast = _fast_path_parse(content, filename, mime, schema_cols)
    if fast is not None:
        final_cols = schema_cols if schema_cols else list(fast.keys())
        source     = schema_source if schema_cols else "file_headers"
        log.info("_parse_expected_output: fast-path success source=%s cols=%s", source, final_cols)
        return fast, False, final_cols, source

    # Step 3: render file as text for LLM
    file_text = _render_file_as_text(content, filename, mime)
    if not file_text.strip():
        log.warning("_parse_expected_output: unreadable file %s — placeholders inserted", filename)
        placeholder = {c: None for c in schema_cols} if schema_cols else {}
        return placeholder, True, schema_cols, schema_source

    # Step 4: LLM mapping (handles synonyms, casing, orientation, no exact schema match)
    llm_result = _llm_map_to_schema(schema_cols, file_text, board_id)
    if llm_result is None:
        placeholder = {c: None for c in schema_cols} if schema_cols else {}
        log.warning("_parse_expected_output: LLM mapping failed — placeholders inserted")
        return placeholder, True, schema_cols, schema_source

    # Step 5: reconcile LLM result with known schema (if any)
    if schema_cols:
        answer_key: dict = {}
        for col in schema_cols:
            # Try exact match first, then case-insensitive fallback
            val = llm_result.get(col)
            if val is None:
                col_lo = col.lower()
                for k, v in llm_result.items():
                    if k.lower() == col_lo:
                        val = v
                        break
            answer_key[col] = val
        final_cols    = schema_cols
        actual_source = schema_source
    else:
        # No board schema — LLM inferred the schema from the file
        answer_key    = llm_result
        final_cols    = list(llm_result.keys())
        actual_source = "file_inferred"

    log.info(
        "_parse_expected_output: LLM-mapped %d fields source=%s",
        len(answer_key), actual_source,
    )
    return answer_key, False, final_cols, actual_source


# ── Storage upload ─────────────────────────────────────────────────────────────

def _upload_to_storage(sb, content: bytes, board_id: str, filename: str, mime: str) -> str:
    """Upload bytes to sample-files bucket. Returns storage_path."""
    safe_name    = filename.replace(" ", "_")
    storage_path = f"{STORAGE_PREFIX}/{board_id}/{safe_name}"
    try:
        try:
            sb.storage.from_(BUCKET).remove([storage_path])
        except Exception:
            pass
        sb.storage.from_(BUCKET).upload(
            path=storage_path,
            file=content,
            file_options={"content-type": mime or "application/octet-stream"},
        )
    except Exception as exc:
        log.warning("Storage upload failed for %s: %s — continuing without storage_path", filename, exc)
        return ""
    return storage_path


# ── Route ──────────────────────────────────────────────────────────────────────

@router.post("/{board_id}/worked-example")
async def capture_worked_example(
    board_id: str,
    email: UploadFile = File(...),
    attachments: List[UploadFile] = File(default=[]),
    expected_output: UploadFile = File(...),
    fixture_subject: Optional[str] = Form(default=""),
) -> dict:
    """Upload a worked example (email + attachments + expected output) for this board.

    Derives the expected output schema from the board spec, then uses an LLM to
    map the uploaded file onto that schema — tolerating synonyms, ordering, and
    column/row orientation.  Idempotent: re-upload replaces the existing case.
    """
    try:
        sb = _sb()

        # ── 1. Clear existing sample_files for this board ─────────────────────
        sb.table("sample_files").delete().eq("board_id", board_id).execute()
        log.info("worked-example: cleared old sample_files for board %s", board_id)

        sample_file_rows: list[dict] = []

        # ── 2. Upload + register the email file ───────────────────────────────
        email_bytes   = await email.read()
        email_name    = email.filename or "sample_email.txt"
        email_mime    = email.content_type or "text/plain"
        email_storage = _upload_to_storage(
            sb, email_bytes, board_id,
            "sample_email" + (pathlib.Path(email_name).suffix or ".txt"),
            email_mime,
        )
        email_text = _extract_text(email_bytes, email_name, email_mime)

        sample_file_rows.append({
            "board_id":       board_id,
            "node_id":        None,
            "filename":       "sample_email" + (pathlib.Path(email_name).suffix or ".txt"),
            "mime":           email_mime,
            "storage_path":   email_storage,
            "extracted_text": email_text or "",
        })
        log.info(
            "worked-example: email → %s (%d chars)",
            sample_file_rows[-1]["filename"], len(email_text or ""),
        )

        # ── 3. Upload + register each attachment ──────────────────────────────
        for att in attachments:
            att_bytes   = await att.read()
            att_name    = att.filename or f"attachment_{uuid.uuid4().hex[:8]}"
            att_mime    = att.content_type or "application/octet-stream"
            att_storage = _upload_to_storage(sb, att_bytes, board_id, att_name, att_mime)
            att_text    = _extract_text(att_bytes, att_name, att_mime)

            sample_file_rows.append({
                "board_id":       board_id,
                "node_id":        None,
                "filename":       att_name,
                "mime":           att_mime,
                "storage_path":   att_storage,
                "extracted_text": att_text or "",
            })
            log.info(
                "worked-example: attachment %s → storage_path=%s text=%d chars",
                att_name, att_storage or "(none)", len(att_text or ""),
            )

        # ── 4. Insert all sample_files rows ───────────────────────────────────
        if sample_file_rows:
            sb.table("sample_files").insert(sample_file_rows).execute()
        log.info("worked-example: inserted %d sample_files rows", len(sample_file_rows))

        # ── 5. Parse expected output → answer key ─────────────────────────────
        eo_bytes  = await expected_output.read()
        eo_name   = expected_output.filename or "expected.csv"
        eo_mime   = expected_output.content_type or "text/csv"

        answer_key, needs_confirmation, output_schema, schema_source = _parse_expected_output(
            eo_bytes, eo_name, eo_mime, board_id, sb,
        )
        log.info(
            "worked-example: parsed %s → needs_confirmation=%s schema_source=%s cols=%s",
            eo_name, needs_confirmation, schema_source, output_schema,
        )

        # ── 6. Write eval case JSON ────────────────────────────────────────────
        CASES_DIR.mkdir(parents=True, exist_ok=True)
        safe_id   = board_id.replace("-", "")[:8]
        case_path = CASES_DIR / f"board_{safe_id}_worked.json"

        case = {
            "_instructions": [
                f"Auto-generated eval case for board {board_id}.",
                f"Output schema derived from: {schema_source}.",
                "Expected values were parsed from the uploaded expected_output file.",
                f"needs_confirmation={needs_confirmation} — set to false once you verify.",
                "Re-run: python -m backend.evals.evaluate " + board_id,
            ],
            "board_id":            board_id,
            "name":                f"Worked example — board {board_id[:8]}",
            "fixture_subject":     (fixture_subject or "").strip(),
            "email_filename_hint": "sample_email",
            "needs_confirmation":  needs_confirmation,
            "output_schema":       output_schema,
            "schema_source":       schema_source,
            "expected":            answer_key,
            "tolerances":          {},
        }
        case_path.write_text(json.dumps(case, indent=2, ensure_ascii=False), encoding="utf-8")
        log.info("worked-example: wrote eval case → %s", case_path)

        return {
            "captured":           True,
            "case_path":          str(case_path),
            "answer_key":         answer_key,
            "needs_confirmation": needs_confirmation,
            "output_schema":      output_schema,
            "schema_source":      schema_source,
            "files_uploaded":     len(sample_file_rows),
            "board_id":           board_id,
        }

    except HTTPException:
        raise
    except Exception as exc:
        log.exception("capture_worked_example failed board_id=%s", board_id)
        raise HTTPException(500, f"worked-example upload failed: {exc}")
